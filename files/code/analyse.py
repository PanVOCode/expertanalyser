import os
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'config'))
from config import (
    only_text, stat_tables, api_id, api_hash, start_date, end_date, SORT_BY_VIRALITY,
    CREATE_MULTIPLE_SORTED_FILES, SHOW_VIRALITY_STATISTICS, CREATE_VIRALITY_SUMMARY_REPORT
)
from telethon.sync import TelegramClient
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import numpy as np
import re
from pathlib import Path

POSITIVE_EMOJIS = {'👍', '❤', '🔥', '😊', '😂', '🥰', '👏', '⚡', '❤‍🔥', '🫡', '🤗', '😍', '👌', '😁', '💯', '🙏', '🤩'}

def calculate_er_percentage(row):
    """Вычисляет процент вовлеченности (ER%)"""
    views = row['Просмотры']
    if views == 0:
        return 0
    
    total_engagement = row['Реакции'] + row['Комменты'] + row['Пересылки']
    return (total_engagement / views) * 100

def calculate_virality(row):
    """Вычисляет виральность поста по основной формуле"""
    forwards = row['Пересылки']
    comments = row['Комменты']
    reactions = row['Реакции']
    er_percentage = row['ER%']
    positive_negative_balance = row['Всего (+)'] - row['Всего (-)']
    views = row['Просмотры']
    
    # Основная формула виральности
    viral_score = (5 * forwards + 3 * comments + 2 * reactions + 1.5 * er_percentage + 0.5 * positive_negative_balance) / np.log(views + 1)
    
    return viral_score

def calculate_viral_coefficient(row):
    """Вычисляет коэффициент виральности (пересылки/просмотры * 1000)"""
    forwards = row['Пересылки']
    views = row['Просмотры']
    
    if views == 0:
        return 0
    
    return (forwards / views) * 1000

def calculate_engagement_virality(row):
    """Вычисляет виральность вовлеченности"""
    reactions = row['Реакции']
    comments = row['Комменты']
    forwards = row['Пересылки']
    views = row['Просмотры']
    
    if views == 0:
        return 0
    
    total_engagement = reactions + comments + forwards
    return (total_engagement / views) * 100

def sort_by_virality(df, sort_type='default'):
    """
    Сортировка по различным метрикам виральности
    
    Args:
        df: DataFrame с данными
        sort_type: тип сортировки ('default', 'coefficient', 'engagement', 'forwards', 'reactions')
    """
    if sort_type == 'default':
        # Сортировка по основной формуле виральности
        return df.sort_values(by='Виральность', ascending=False).reset_index(drop=True)
    elif sort_type == 'coefficient':
        # Сортировка по коэффициенту виральности (пересылки/просмотры)
        return df.sort_values(by='Коэффициент виральности', ascending=False).reset_index(drop=True)
    elif sort_type == 'engagement':
        # Сортировка по виральности вовлеченности
        return df.sort_values(by='Виральность вовлеченности', ascending=False).reset_index(drop=True)
    elif sort_type == 'forwards':
        # Сортировка по количеству пересылок
        return df.sort_values(by='Пересылки', ascending=False).reset_index(drop=True)
    elif sort_type == 'reactions':
        # Сортировка по количеству реакций
        return df.sort_values(by='Реакции', ascending=False).reset_index(drop=True)
    elif sort_type == 'views':
        # Сортировка по просмотрам
        return df.sort_values(by='Просмотры', ascending=False).reset_index(drop=True)
    elif sort_type == 'date':
        # Сортировка по дате (новые сначала)
        return df.sort_values(by='Дата', ascending=False).reset_index(drop=True)
    else:
        # По умолчанию сортировка по основной виральности
        return df.sort_values(by='Виральность', ascending=False).reset_index(drop=True)

def save_to_excel(df, filename, channel_username, start_date, end_date):
    """Сохраняет DataFrame в Excel файл с форматированием"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"
    
    # Добавляем данные
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Форматирование заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Закрепляем заголовки
    ws.freeze_panes = "A2"
    
    # Добавляем информацию о канале и периоде
    ws.insert_rows(1, 3)
    ws['A1'] = f"Канал: {channel_username}"
    ws['A2'] = f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    ws['A3'] = f"Всего постов: {len(df)}"
    
    # Сохраняем файл в папку results
    results_dir = Path('results')
    results_dir.mkdir(exist_ok=True)
    file_path = results_dir / filename
    wb.save(file_path)
    print(f"Файл сохранен: {file_path}")

def create_multiple_sorted_files(df, channel_username, start_date, end_date):
    """
    Создает несколько файлов Excel с разными типами сортировки по виральности
    """
    # Исключаем строки с нечисловыми значениями в ключевых столбцах
    df = df[~df['Дата'].isin(['Итого', 'В среднем на пост'])].copy()
    # Приводим все числовые столбцы к числовому типу
    numeric_columns = ['Просмотры', 'Реакции', 'Всего (+)', 'Всего (-)', 'Всего', 'Комменты', 'Пересылки', 'ER%', 'Виральность', 'Коэффициент виральности', 'Виральность вовлеченности']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    sort_types = {
        'default': 'По виральности (основная формула)',
        'coefficient': 'По коэффициенту виральности',
        'engagement': 'По виральности вовлеченности',
        'forwards': 'По пересылкам',
        'reactions': 'По реакциям',
        'views': 'По просмотрам',
        'date': 'По дате'
    }
    
    for sort_type, description in sort_types.items():
        if not stat_tables.get(sort_type, False):
            continue
        sorted_df = sort_by_virality(df.copy(), sort_type)
        filename = f'posts_sorted_by_{sort_type}.xlsx'
        save_to_excel(sorted_df, filename, channel_username, start_date, end_date)
        print(f"Создан файл: {filename} - {description}")

def print_virality_statistics(df):
    """
    Выводит статистику по виральности постов
    """
    # Исключаем строки с нечисловыми значениями в ключевых столбцах
    df = df[~df['Дата'].isin(['Итого', 'В среднем на пост'])].copy()
    # Приводим все числовые столбцы к числовому типу
    numeric_columns = ['Просмотры', 'Реакции', 'Всего (+)', 'Всего (-)', 'Всего', 'Комменты', 'Пересылки', 'ER%', 'Виральность', 'Коэффициент виральности', 'Виральность вовлеченности']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    print("\n" + "="*60)
    print("СТАТИСТИКА ПО ВИРАЛЬНОСТИ")
    print("="*60)
    
    # Общая статистика
    print(f"Всего постов: {len(df)}")
    print(f"Средняя виральность: {df['Виральность'].mean():.2f}")
    print(f"Максимальная виральность: {df['Виральность'].max():.2f}")
    print(f"Минимальная виральность: {df['Виральность'].min():.2f}")
    
    # Дополнительная статистика
    print(f"\nОбщая статистика:")
    print(f"Средний ER%: {df['ER%'].mean():.2f}%")
    print(f"Среднее количество комментариев: {df['Комменты'].mean():.1f}")
    print(f"Общее количество комментариев: {df['Комменты'].sum():,}")
    print(f"Средний баланс реакций (+/-): {df['Всего (+)'].mean() - df['Всего (-)'].mean():.1f}")
    print(f"Общий баланс реакций (+/-): {df['Всего (+)'].sum() - df['Всего (-)'].sum():,}")
    
    # Top-5 sections
    if stat_tables.get('default', False):
        print("\nТОП-5 ПОСТОВ ПО ВИРАЛЬНОСТИ:")
        print("-" * 40)
        top_viral = df.nlargest(5, 'Виральность')
        for i, (_, row) in enumerate(top_viral.iterrows(), 1):
            print(f"{i}. Виральность: {row['Виральность']:.2f}")
            print(f"   Дата: {row['Дата']}")
            print(f"   Просмотры: {row['Просмотры']:,}")
            print(f"   Пересылки: {row['Пересылки']}")
            print(f"   Реакции: {row['Реакции']}")
            print(f"   Комментарии: {row['Комменты']}")
            print(f"   ER%: {row['ER%']:.2f}%")
            print(f"   Баланс реакций (+/-): {row['Всего (+)'] - row['Всего (-)']}")
            print(f"   Тип: {row['Тип']}")
            print()
    
    if stat_tables.get('coefficient', False):
        print("\nТОП-5 ПОСТОВ ПО КОЭФФИЦИЕНТУ ВИРАЛЬНОСТИ:")
        print("-" * 40)
        top_coefficient = df.nlargest(5, 'Коэффициент виральности')
        for i, (_, row) in enumerate(top_coefficient.iterrows(), 1):
            print(f"{i}. Коэффициент: {row['Коэффициент виральности']:.2f}")
            print(f"   Дата: {row['Дата']}")
            print(f"   Пересылки: {row['Пересылки']}")
            print(f"   Просмотры: {row['Просмотры']:,}")
            print(f"   Тип: {row['Тип']}")
            print()
    
    if stat_tables.get('engagement', False):
        print("\nТОП-5 ПОСТОВ ПО ВИРАЛЬНОСТИ ВОВЛЕЧЕННОСТИ:")
        print("-" * 40)
        top_engagement = df.nlargest(5, 'Виральность вовлеченности')
        for i, (_, row) in enumerate(top_engagement.iterrows(), 1):
            print(f"{i}. Виральность вовлеченности: {row['Виральность вовлеченности']:.2f}%")
            print(f"   Дата: {row['Дата']}")
            print(f"   Общее вовлечение: {row['Реакции'] + row['Комменты'] + row['Пересылки']}")
            print(f"   Просмотры: {row['Просмотры']:,}")
            print(f"   Тип: {row['Тип']}")
            print()
    
    if stat_tables.get('forwards', False):
        print("\nТОП-5 ПОСТОВ ПО ПЕРЕСЫЛКАМ:")
        print("-" * 40)
        top_forwards = df.nlargest(5, 'Пересылки')
        for i, (_, row) in enumerate(top_forwards.iterrows(), 1):
            print(f"{i}. Пересылки: {row['Пересылки']}")
            print(f"   Дата: {row['Дата']}")
            print(f"   Просмотры: {row['Просмотры']:,}")
            print(f"   Тип: {row['Тип']}")
            print()
    
    if stat_tables.get('reactions', False):
        print("\nТОП-5 ПОСТОВ ПО РЕАКЦИЯМ:")
        print("-" * 40)
        top_reactions = df.nlargest(5, 'Реакции')
        for i, (_, row) in enumerate(top_reactions.iterrows(), 1):
            print(f"{i}. Реакции: {row['Реакции']}")
            print(f"   Дата: {row['Дата']}")
            print(f"   Просмотры: {row['Просмотры']:,}")
            print(f"   Тип: {row['Тип']}")
            print()
    
    if stat_tables.get('views', False):
        print("\nТОП-5 ПОСТОВ ПО ПРОСМОТРАМ:")
        print("-" * 40)
        top_views = df.nlargest(5, 'Просмотры')
        for i, (_, row) in enumerate(top_views.iterrows(), 1):
            print(f"{i}. Просмотры: {row['Просмотры']:,}")
            print(f"   Дата: {row['Дата']}")
            print(f"   Пересылки: {row['Пересылки']}")
            print(f"   Реакции: {row['Реакции']}")
            print(f"   Тип: {row['Тип']}")
            print()

def create_virality_summary_report(df, channel_username, start_date, end_date):
    """
    Создает сводный отчет по виральности в отдельном файле
    """
    # Исключаем строки с нечисловыми значениями в ключевых столбцах
    df = df[~df['Дата'].isin(['Итого', 'В среднем на пост'])].copy()
    # Приводим все числовые столбцы к числовому типу
    numeric_columns = ['Просмотры', 'Реакции', 'Всего (+)', 'Всего (-)', 'Всего', 'Комменты', 'Пересылки', 'ER%', 'Виральность', 'Коэффициент виральности', 'Виральность вовлеченности']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    report_data = []
    
    # Общая статистика
    report_data.append({
        'Метрика': 'Всего постов',
        'Значение': len(df),
        'Описание': 'Общее количество проанализированных постов'
    })
    
    report_data.append({
        'Метрика': 'Средняя виральность',
        'Значение': f"{df['Виральность'].mean():.2f}",
        'Описание': 'Среднее значение виральности по всем постам'
    })
    
    report_data.append({
        'Метрика': 'Максимальная виральность',
        'Значение': f"{df['Виральность'].max():.2f}",
        'Описание': 'Максимальное значение виральности'
    })
    
    report_data.append({
        'Метрика': 'Минимальная виральность',
        'Значение': f"{df['Виральность'].min():.2f}",
        'Описание': 'Минимальное значение виральности'
    })
    
    # Дополнительная статистика
    report_data.append({
        'Метрика': 'Средний ER%',
        'Значение': f"{df['ER%'].mean():.2f}%",
        'Описание': 'Средний процент вовлеченности'
    })
    
    report_data.append({
        'Метрика': 'Среднее количество комментариев',
        'Значение': f"{df['Комменты'].mean():.1f}",
        'Описание': 'Среднее количество комментариев на пост'
    })
    
    report_data.append({
        'Метрика': 'Общее количество комментариев',
        'Значение': f"{df['Комменты'].sum():,}",
        'Описание': 'Общее количество комментариев по всем постам'
    })
    
    report_data.append({
        'Метрика': 'Средний баланс реакций (+/-)',
        'Значение': f"{df['Всего (+)'].mean() - df['Всего (-)'].mean():.1f}",
        'Описание': 'Средний баланс позитивных и негативных реакций'
    })
    
    report_data.append({
        'Метрика': 'Общий баланс реакций (+/-)',
        'Значение': f"{df['Всего (+)'].sum() - df['Всего (-)'].sum():,}",
        'Описание': 'Общий баланс позитивных и негативных реакций'
    })
    
    # Топ посты
    top_viral = df.nlargest(3, 'Виральность')
    for i, (_, row) in enumerate(top_viral.iterrows(), 1):
        report_data.append({
            'Метрика': f"{i}. {row['Дата']}",
            'Значение': f"Виральность: {row['Виральность']:.2f}",
            'Описание': f"Просмотры: {row['Просмотры']:,}, Пересылки: {row['Пересылки']}, Комментарии: {row['Комменты']}, ER%: {row['ER%']:.2f}%, Баланс: {row['Всего (+)'] - row['Всего (-)']}, Тип: {row['Тип']}"
        })
    
    # Статистика по типам контента
    content_stats = df.groupby('Тип').agg({
        'Виральность': 'mean',
        'Просмотры': 'mean',
        'Пересылки': 'mean',
        'Реакции': 'mean',
        'Комменты': 'mean'
    }).round(2)
    
    for content_type, stats in content_stats.iterrows():
        report_data.append({
            'Метрика': f"Тип: {content_type}",
            'Значение': f"Средняя виральность: {stats['Виральность']:.2f}",
            'Описание': f"Средние показатели - Просмотры: {stats['Просмотры']:.0f}, Пересылки: {stats['Пересылки']:.1f}, Реакции: {stats['Реакции']:.1f}, Комментарии: {stats['Комменты']:.1f}"
        })
    
    # Создаем DataFrame для отчета
    report_df = pd.DataFrame(report_data)
    
    # Сохраняем отчет
    filename = 'virality_summary_report.xlsx'
    save_to_excel(report_df, filename, channel_username, start_date, end_date)
    print(f"Создан сводный отчет: {filename}")

def analyse(channel_url):
    """
    Основная функция анализа канала Telegram
    
    Args:
        channel_url (str): URL канала для анализа (например, "https://t.me/sellerx")
        
    Returns:
        str: Текст из топ-5 постов по виральности в формате "текст 1: текст\nтекст 2: текст\n..."
    """
    # Импортируем функции для использования внутри analyse
    from analyse import save_to_excel, create_multiple_sorted_files, print_virality_statistics, create_virality_summary_report
    
    # Извлекаем username из ссылки
    match = re.search(r"(?:https?://)?t\.me/(.+?)(?:/|$)", channel_url)
    if not match:
        raise ValueError("Некорректная ссылка на канал")
    channel_username = match.group(1).replace('@', '')
    
    # Создаем папку для результатов в results/all_folders
    out_dir = Path('results', 'all_folders') / channel_username
    out_dir.mkdir(parents=True, exist_ok=True)
    
    with TelegramClient('session_name', api_id, api_hash) as client:
        messages = []
        albums = {}
        
        try:
            # Получаем канал
            channel = client.get_entity(channel_username)
            print(f"Анализируем канал: {channel.title}")
            
            # Получаем сообщения за указанный период
            for message in client.iter_messages(channel, offset_date=start_date, reverse=True):
                if message.date > end_date:
                    break
                
                # Пропускаем служебные сообщения
                if message.action:
                    continue
                
                # Определяем тип контента
                content_type = "Текст"
                if message.photo:
                    content_type = "Фото"
                elif message.video:
                    content_type = "Видео"
                elif message.document:
                    content_type = "Документ"
                elif message.voice:
                    content_type = "Голосовое"
                elif message.video_note:
                    content_type = "Видеосообщение"
                elif message.sticker:
                    content_type = "Стикер"
                elif message.poll:
                    content_type = "Опрос"
                
                # Получаем реакции
                reactions = message.reactions
                positive_reactions = 0
                negative_reactions = 0
                total_reactions = 0
                
                if reactions:
                    try:
                        # Правильный способ доступа к реакциям в Telethon
                        for reaction in reactions.results:
                            emoji = reaction.reaction.emoji if hasattr(reaction.reaction, 'emoji') else '👍'
                            count = reaction.count
                            total_reactions += count
                            
                            if emoji in POSITIVE_EMOJIS:
                                positive_reactions += count
                            else:
                                negative_reactions += count
                    except AttributeError:
                        # Альтернативный способ для старых версий Telethon
                        try:
                            for reaction in reactions.reactions:
                                emoji = reaction.emoji
                                count = reaction.count
                                total_reactions += count
                                
                                if emoji in POSITIVE_EMOJIS:
                                    positive_reactions += count
                                else:
                                    negative_reactions += count
                        except:
                            # Если не удается получить реакции, оставляем нули
                            pass
                
                # Получаем количество комментариев
                comments_count = 0
                try:
                    # Пытаемся получить количество комментариев
                    if hasattr(message, 'replies') and message.replies:
                        comments_count = message.replies.replies
                except:
                    pass
                
                # Получаем количество пересылок
                forwards_count = 0
                try:
                    if hasattr(message, 'forwards'):
                        forwards_count = message.forwards
                except:
                    pass
                
                # Получаем количество просмотров
                views_count = 0
                try:
                    if hasattr(message, 'views'):
                        views_count = message.views
                except:
                    pass
                
                message_data = {
                    'Дата': message.date.strftime('%d.%m.%Y %H:%M'),
                    'Полный_текст': message.text if message.text else '',
                    'Тип': content_type,
                    'Просмотры': views_count,
                    'Реакции': total_reactions,
                    'Всего (+)': positive_reactions,
                    'Всего (-)': negative_reactions,
                    'Всего': total_reactions,
                    'Комменты': comments_count,
                    'Пересылки': forwards_count
                }
                
                messages.append(message_data)
                
                # Обработка альбомов
                if message.grouped_id:
                    if message.grouped_id not in albums:
                        albums[message.grouped_id] = []
                    albums[message.grouped_id].append(message_data)
        
        except Exception as e:
            print(f"Ошибка при получении сообщений: {e}")
            print(f"Тип ошибки: {type(e).__name__}")
            return ""
        
        print(f"Всего получено сообщений: {len(messages)}")
        
        if not messages:
            print("Сообщения не найдены")
            return ""
        
        # Создаем DataFrame
        df = pd.DataFrame(messages)
        
        # Фильтрация по only_text
        if only_text:
            df = df[df['Тип'] == 'Текст'].reset_index(drop=True)
        
        df = df.sort_values(by='Дата', ascending=False).reset_index(drop=True)
        
        # Приводим числовые столбцы к числовому типу
        numeric_columns = ['Просмотры', 'Реакции', 'Всего (+)', 'Всего (-)', 'Всего', 'Комменты', 'Пересылки']
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # Вычисляем ER%
        df['ER%'] = df.apply(calculate_er_percentage, axis=1)
        
        # Вычисляем виральность
        df['Виральность'] = df.apply(calculate_virality, axis=1)
        df['Коэффициент виральности'] = df.apply(calculate_viral_coefficient, axis=1)
        df['Виральность вовлеченности'] = df.apply(calculate_engagement_virality, axis=1)
        
        # Сортировка по выбранному типу виральности
        df = sort_by_virality(df, SORT_BY_VIRALITY)
        
        # Добавляем итоговую строку
        total_row = {
            'Дата': 'Итого',
            'Полный_текст': '',
            'Тип': '',
            'Просмотры': df['Просмотры'].sum(),
            'Реакции': df['Реакции'].sum(),
            'Всего (+)': df['Всего (+)'].sum(),
            'Всего (-)': df['Всего (-)'].sum(),
            'Всего': df['Всего'].sum(),
            'Комменты': df['Комменты'].sum(),
            'Пересылки': df['Пересылки'].sum(),
            'ER%': (df['Реакции'].sum() + df['Комменты'].sum() + df['Пересылки'].sum()) / df['Просмотры'].sum() * 100 if df['Просмотры'].sum() > 0 else 0,
            'Виральность': df['Виральность'].sum(),
            'Коэффициент виральности': df['Коэффициент виральности'].sum(),
            'Виральность вовлеченности': df['Виральность вовлеченности'].sum()
        }
        
        # Добавляем строку со средними значениями
        avg_row = {
            'Дата': 'В среднем на пост',
            'Полный_текст': '',
            'Тип': '',
            'Просмотры': df['Просмотры'].mean(),
            'Реакции': df['Реакции'].mean(),
            'Всего (+)': df['Всего (+)'].mean(),
            'Всего (-)': df['Всего (-)'].mean(),
            'Всего': df['Всего'].mean(),
            'Комменты': df['Комменты'].mean(),
            'Пересылки': df['Пересылки'].mean(),
            'ER%': df['ER%'].mean(),
            'Виральность': df['Виральность'].mean(),
            'Коэффициент виральности': df['Коэффициент виральности'].mean(),
            'Виральность вовлеченности': df['Виральность вовлеченности'].mean()
        }
        
        # Добавляем итоговые строки
        df = pd.concat([df, pd.DataFrame([total_row, avg_row])], ignore_index=True)
        
        # Сохраняем основной файл
        save_to_excel(df, 'posts.xlsx', channel_username, start_date, end_date)
        
        # Создание дополнительных файлов с разными типами сортировки
        if CREATE_MULTIPLE_SORTED_FILES:
            print("\nСоздание файлов с разными типами сортировки по виральности...")
            # Переопределяем save_to_excel для сохранения в out_dir (temporary override)
            def save_to_excel_dir(df, filename, channel_username, start_date, end_date):
                # Сохраняем в папку канала в results/all_folders
                file_path = out_dir / filename
                # Создаем временную функцию для сохранения в конкретную папку
                def save_to_excel_temp(df, filename, channel_username, start_date, end_date):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Posts"
                    
                    # Добавляем данные
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    
                    # Форматирование заголовков
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Автоматическая ширина столбцов
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Закрепляем заголовки
                    ws.freeze_panes = "A2"
                    
                    # Добавляем информацию о канале и периоде
                    ws.insert_rows(1, 3)
                    ws['A1'] = f"Канал: {channel_username}"
                    ws['A2'] = f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
                    ws['A3'] = f"Всего постов: {len(df)}"
                    
                    # Сохраняем файл
                    wb.save(file_path)
                    print(f"Файл сохранен: {file_path}")
                
                save_to_excel_temp(df, filename, channel_username, start_date, end_date)
            
            # Сохраняем оригинальную функцию и заменяем её
            old_save_to_excel = save_to_excel
            save_to_excel = save_to_excel_dir
            create_multiple_sorted_files(df, channel_username, start_date, end_date)
            save_to_excel = old_save_to_excel  # Restore original save_to_excel
            print("Все файлы созданы успешно!")
        
        # Вывод статистики по виральности
        if SHOW_VIRALITY_STATISTICS:
            print_virality_statistics(df)
        
        # Создание сводного отчета по виральности
        if CREATE_VIRALITY_SUMMARY_REPORT:
            create_virality_summary_report(df, channel_username, start_date, end_date)
        
        # Формируем строку с текстами из топ-5 постов
        top_posts_text = ""
        
        # Исключаем итоговые строки для получения топ-5
        df_for_top = df[~df['Дата'].isin(['Итого', 'В среднем на пост'])].copy()
        
        if len(df_for_top) > 0:
            # Получаем топ-5 по виральности
            top_5_posts = df_for_top.nlargest(5, 'Виральность')
            
            for i, (_, row) in enumerate(top_5_posts.iterrows(), 1):
                post_text = row.get('Полный_текст', '')
                if post_text and post_text.strip():
                    top_posts_text += f"текст {i}:\n{post_text.strip()}\n\n"
        
        return top_posts_text 